import numpy as np
import pandas as pd

import openpyxl
import random
from itertools import combinations


def main():
    n = 5

    # 读取10个执行体
    _data = pd.read_excel('C:/Users/user/Desktop/demo.xlsx',
                          usecols=['序号', 'd1', 'd2', 'd3', '异构度', "运行开销", '切换开销', "选中次数", '置信度', '执行体类型', '工作时长'])
    data = np.array(_data)
    print("执行体池：")
    print("序号", "代码相异值", "传输相异值", "运行相异值", '异构度', "运行开销", '切换开销', "选中次数", '置信度', '执行体类型', '工作时长')
    print(data)

    # 初始时从10个执行体中随机选择5个作为服务体
    print("初始时, 从10个异构体中随机选取5个组成服务体S")
    s = random.sample(range(1, 11), n)
    print(s)
    print("初始平均置信度H(t):0.5")

    # print("服务体S：")
    # print("序号", "代码相异值", "传输相异值", "运行相异值", '异构度', "运行开销", '切换开销', "选中次数", '置信度', '执行体类型')
    # 打印服务体
    _E = [data[s[0] - 1], data[s[1] - 1], data[s[2] - 1], data[s[3] - 1], data[s[4] - 1]]
    E = np.array(_E)
    # print(E)

    # 标志位，用于第一次打开test文件，后续在test_2文件操作
    flag = 0

    for var in range(1, 101):
        print("-------------------------------第" + str(var) + "次迭代-------------------------------")
        print("服务体S：")
        print("序号", "代码相异值", "传输相异值", "运行相异值", '异构度', "运行开销", '切换开销', "选中次数", '置信度', '执行体类型', '工作时长')
        # 执行体选中次数+1
        E[0][7] += 1
        E[1][7] += 1
        E[2][7] += 1
        E[3][7] += 1
        E[4][7] += 1
        print(E)

        # 存放被选中的执行体的序号
        ids = [int(E[0][0]), int(E[1][0]), int(E[2][0]), int(E[3][0]), int(E[4][0])]
        # print(ids)

        # 0.5-1.5之间随机数 保留1位
        # x = '{:.1f}'.format(random.uniform(0.5, 1.5))

        # 字典y存放执行体局部裁决结果
        y = {}
        for i in range(0, n):
            if E[i][9] == 1:
                x = '{:.1f}'.format(random.uniform(0.5, 1.5))
            else:
                x = '{:.1f}'.format(random.uniform(0.0, 2.0))
            y[int(E[i][0])] = x

        print("局部裁决结果：")
        print(y)

        # sum_h代表选中的执行体的置信度之和，sum_hy代表各个执行体置信度,Y代表加权平均值
        sum_h = 0.0
        sum_hy = 0.0
        for i in range(0, n):
            sum_h += E[i][8]
            sum_hy += E[i][8] * float(y[E[i][0]])
        # print(sum_h)
        # print(sum_hy)
        Y = '{:.2f}'.format(sum_hy / sum_h)
        Y = float(Y)
        print("加权平均值：")
        print(Y)

        # 归一化输出差异,并根据误差阈值，相应调整执行体的置信度
        print("归一化输出差异：")
        for i in range(0, n):
            print(abs((float(y[E[i][0]]) - Y) / Y))
            if abs((float(y[E[i][0]]) - Y) / Y) <= 0.5:
                E[i][8] = ((E[i][7]) * E[i][8] + 1) / (E[i][7] + 1)
            else:
                E[i][8] = ((E[i][7]) * E[i][8]) / (E[i][7] + 1)

        print("动态更新执行体置信度：")
        for i in range(0, n):
            print(E[i][8])

        print("服务体S：")
        print("序号", "代码相异值", "传输相异值", "运行相异值", '异构度', "运行开销", '切换开销', "选中次数", '置信度', '执行体类型', '工作时长')
        print(E)

        H = 0
        for i in range(0, n):
            H += E[i][8]
        H = H / n
        print("第" + str(var) + "次迭代后，服务体的平均信誉度H(t)：" + str(H))

        # 存储平均置信度
        wb = openpyxl.load_workbook('n5_confidence.xlsx')
        ws = wb['Sheet']

        ws.cell(row=var + 1, column=1).value = var
        ws.cell(row=var + 1, column=2).value = H

        wb.save('n5_confidence.xlsx')

        D = 0
        for i in range(0, 5):
            D += E[i][4]
        D = D / 5
        print("第" + str(var) + "次迭代后，服务体的平均异构度D(t)：" + str(D))

        # 存储异构度
        wb = openpyxl.load_workbook('n5_D.xlsx')
        ws = wb['Sheet']

        ws.cell(row=var + 1, column=1).value = var
        ws.cell(row=var + 1, column=2).value = D

        wb.save('n5_D.xlsx')

        # 紧急切换准备工作
        # 紧急切换，系统置信度阈值初设0.25
        # error记录置信度低于阈值的服务体的个数
        error = []
        for i in range(0, n):
            if E[i][8] < 0.25:
                error.append(i)
        if len(error) == 1:
            print("序号为" + str(E[error[0]][0]) + "的服务体置信度低于系统置信度阈值，触发紧急切换！")
            # 清洗（将置信度置为初值0.5）
            E[error[0]][8] = 0.5
        elif len(error) == 2:
            print("序号为" + str(E[error[0]][0]) + "和" + str(E[error[1]][0]) + "的服务体置信度低于系统置信度阈值，触发紧急切换！")
            # 清洗（将置信度置为初值0.5）
            E[error[0]][8] = 0.5
            E[error[1]][8] = 0.5
        elif len(error) == 3:
            print("序号为" + str(E[error[0]][0]) + "、" + str(E[error[1]][0]) + "和" + str(E[error[2]][0]) + "的服务体置信度低于系统置信度阈值，触发紧急切换！")
            # 清洗（将置信度置为初值0.5）
            E[error[0]][8] = 0.5
            E[error[1]][8] = 0.5
            E[error[2]][8] = 0.5
        elif len(error) == 4:
            print("序号为" + str(E[error[0]][0]) + "、" + str(E[error[1]][0]) + "、" + str(E[error[2]][0]) + "和" + str(
                E[error[3]][0]) + "的服务体置信度低于系统置信度阈值，触发紧急切换！")
            # 清洗（将置信度置为初值0.5）
            E[error[0]][8] = 0.5
            E[error[1]][8] = 0.5
            E[error[2]][8] = 0.5
            E[error[3]][8] = 0.5
        elif len(error) == 5:
            print("序号为" + str(E[error[0]][0]) + "、" + str(E[error[1]][0]) + "、" + str(E[error[2]][0]) + "、" + str(E[error[3]][0]) + "和" + str(
                E[error[4]][0]) + "的服务体置信度低于系统置信度阈值，触发紧急切换！")
            # 清洗（将置信度置为初值0.5）
            E[error[0]][8] = 0.5
            E[error[1]][8] = 0.5
            E[error[2]][8] = 0.5
            E[error[3]][8] = 0.5
            E[error[4]][8] = 0.5
        else:
            print("服务体置信度均满足系统阈值！")

        # 将结果动态更新至表格test_2中

        if flag == 0:
            wb = openpyxl.load_workbook('C:/Users/user/Desktop/demo.xlsx')
            flag = 1
        else:
            wb = openpyxl.load_workbook('C:/Users/user/Desktop/n5.xlsx')
        ws = wb['异构度生成']

        for rowNum in range(2, ws.max_row + 1):
            id_ = ws.cell(row=rowNum, column=1).value
            for i in range(0, len(ids)):
                if id_ == ids[i]:
                    ws.cell(row=rowNum, column=11).value = E[i][7]
                    ws.cell(row=rowNum, column=12).value = E[i][8]
                else:
                    i += 1

        wb.save('C:/Users/user/Desktop/n5.xlsx')

        _data_new = pd.read_excel('C:/Users/user/Desktop/n5.xlsx',
                                  usecols=['序号', 'd1', 'd2', 'd3', '异构度', "运行开销", '切换开销', "选中次数", '置信度',
                                           '执行体类型', '工作时长'])
        data_new = np.array(_data_new)

        # Benefit服务质量，H平均置信度，D平均异构度,C_run运行开销,C_switch切换开销
        Benefits = []
        H = []
        D = []
        C_run = []
        C_switch = []

        word = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
        res = [i for i in combinations(word, n)]

        # 紧急切换
        if len(error) != 0:
            print("===============================紧急切换===============================")
            print("执行体池：")
            print("序号", "代码相异值", "传输相异值", "运行相异值", '异构度', "运行开销", '切换开销', "选中次数", '置信度', '执行体类型', '工作时长')
            print(data_new)
            for i in range(0, 252):
                H.append((data_new[res[i][0]][8] + data_new[res[i][1]][8] + data_new[res[i][2]][8] + data_new[res[i][3]][8] + data_new[res[i][4]][8]) / n)
                D.append((data_new[res[i][0]][4] + data_new[res[i][1]][4] + data_new[res[i][2]][4] + data_new[res[i][3]][4] + data_new[res[i][4]][4]) / n)
                C_run.append((data_new[res[i][0]][5] + data_new[res[i][1]][5] + data_new[res[i][2]][5] + data_new[res[i][3]][5] + data_new[res[i][4]][5]))
                Benefits.append(H[i] * D[i] * n / C_run[i])
                # 计算切换开销
                switch = ids.copy()
                c_switch = 0
                for j in range(0, 5):
                    if (res[i][j] + 1) in ids:
                        switch.remove(res[i][j] + 1)
                    else:
                        switch.append(res[i][j] + 1)
                for swi in switch:
                    c_switch += data_new[swi - 1][6]
                C_switch.append(c_switch)

            print("服务质量Benefits:")
            print(Benefits)
            print("切换开销C_switch:")
            print(C_switch)

            # 服务质量（差值计算）
            Benefits_nor = []

            H_now = (data_new[ids[0] - 1][8] + data_new[ids[1] - 1][8] + data_new[ids[2] - 1][8] + data_new[ids[3] - 1][8] + data_new[ids[4] - 1][8]) / n
            D_now = (data_new[ids[0] - 1][4] + data_new[ids[1] - 1][4] + data_new[ids[2] - 1][4] + data_new[ids[3] - 1][4] + data_new[ids[4] - 1][4]) / n
            C_run_now = (data_new[ids[0] - 1][5] + data_new[ids[1] - 1][5] + data_new[ids[2] - 1][5] + data_new[ids[3] - 1][5] + data_new[ids[4] - 1][5])
            Benefits_now = H_now * D_now * n / C_run_now

            for i in range(0, 252):
                Benefits_nor.append(abs((Benefits_now - Benefits[i]) / (Benefits_now + Benefits[i])))

            print("服务质量Benefits_nor(差值计算):")
            print(Benefits_nor)

            C_switch_max = max(C_switch)
            C_switch_min = min(C_switch)

            # print(C_switch_max)
            # print(C_switch_min)
            # 切换开销（归一化后）
            C_switch_nor = []
            for i in range(0, 252):
                C_switch_nor.append((C_switch[i] - C_switch_min) / (C_switch_max - C_switch_min))

            print("切换开销C_switch_nor(归一化后):")
            print(C_switch_nor)

            # 字典存放对应的调度方案
            words = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
            ret = [i for i in combinations(words, 5)]
            a = {}
            for i in range(0, 252):
                a[i] = ret[i]

            # B代表整体收益，Plan_A代表首个满足约束的调度方案
            B = {}
            Plan_A = []

            for i in range(0, 252):
                if Benefits_nor[i] >= 0.3 and C_switch_nor[i] <= 0.7:
                    # C_switch 不能为0
                    if C_switch[i] != 0:
                        B[i] = Benefits[i] / 2 + 1 / (2 * C_switch[i])

            if len(B) != 0:
                print("紧急调度方案：")
                Plan_A.append(a.get((get_dict_key(B, list(B.values())[0]))))

                # 存储系统效益
                wb = openpyxl.load_workbook('n5_Benefit.xlsx')
                ws = wb['Sheet']

                ws.cell(row=var + 1, column=1).value = var
                ws.cell(row=var + 1, column=2).value = list(B.values())[0]

                wb.save('n5_Benefit.xlsx')

                print(Plan_A)
                _E = [data_new[Plan_A[0][0] - 1], data_new[Plan_A[0][1] - 1], data_new[Plan_A[0][2] - 1], data_new[Plan_A[0][3] - 1], data_new[Plan_A[0][4] - 1]]
                E = np.array(_E)
            else:
                print("无满足约束条件的调度方案，清洗后按原方案调度！")

        # 例行切换,假设10个时间长度切换一次
        if var % 10 == 0:
            print("===============================例行切换===============================")
            print("执行体池：")
            print("序号", "代码相异值", "传输相异值", "运行相异值", '异构度', "运行开销", '切换开销', "选中次数", '置信度', '执行体类型', '工作时长')
            print(data_new)
            for i in range(0, 252):
                H.append((data_new[res[i][0]][8] + data_new[res[i][1]][8] + data_new[res[i][2]][8] + data_new[res[i][3]][8] + data_new[res[i][4]][8]) / n)
                D.append((data_new[res[i][0]][4] + data_new[res[i][1]][4] + data_new[res[i][2]][4] + data_new[res[i][3]][4] + data_new[res[i][4]][4]) / n)
                C_run.append((data_new[res[i][0]][5] + data_new[res[i][1]][5] + data_new[res[i][2]][5] + data_new[res[i][3]][5] + data_new[res[i][4]][5]))
                Benefits.append(H[i] * D[i] * n / C_run[i])
                # 计算切换开销
                switch = ids.copy()
                c_switch = 0
                for j in range(0, 5):
                    if (res[i][j] + 1) in ids:
                        switch.remove(res[i][j] + 1)
                    else:
                        switch.append(res[i][j] + 1)
                for swi in switch:
                    c_switch += data_new[swi - 1][6]
                C_switch.append(c_switch)

            print("服务质量Benefits:")
            print(Benefits)
            print("切换开销C_switch:")
            print(C_switch)

            # Benefits_max = max(Benefits)
            # Benefits_min = min(Benefits)

            # 服务质量（差值计算）
            Benefits_nor = []

            H_now = (data_new[ids[0] - 1][8] + data_new[ids[1] - 1][8] + data_new[ids[2] - 1][8] + data_new[ids[3] - 1][8] + data_new[ids[4] - 1][8]) / n
            D_now = (data_new[ids[0] - 1][4] + data_new[ids[1] - 1][4] + data_new[ids[2] - 1][4] + data_new[ids[3] - 1][4] + data_new[ids[4] - 1][4]) / n
            C_run_now = (data_new[ids[0] - 1][5] + data_new[ids[1] - 1][5] + data_new[ids[2] - 1][5] + data_new[ids[3] - 1][5] + data_new[ids[4] - 1][5])
            Benefits_now = H_now * D_now * n / C_run_now

            for i in range(0, 252):
                Benefits_nor.append(abs((Benefits_now - Benefits[i]) / (Benefits_now + Benefits[i])))

            print("服务质量Benefits_nor(差值计算):")
            print(Benefits_nor)

            C_switch_max = max(C_switch)
            C_switch_min = min(C_switch)

            # print(C_switch_max)
            # print(C_switch_min)
            # 切换开销（归一化后）
            C_switch_nor = []
            for i in range(0, 252):
                C_switch_nor.append((C_switch[i] - C_switch_min) / (C_switch_max - C_switch_min))

            print("切换开销C_switch_nor(归一化后):")
            print(C_switch_nor)

            # 字典存放对应的调度方案
            words = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
            ret = [i for i in combinations(words, 5)]
            a = {}
            for i in range(0, 252):
                a[i] = ret[i]

            # B代表整体收益，Plan_B代表最佳调度方案
            B = {}
            Plan_B = []

            for i in range(0, 252):
                if Benefits_nor[i] >= 0.3 and C_switch_nor[i] <= 0.7:
                    # C_switch 不能为0
                    if C_switch[i] != 0:
                        B[i] = Benefits[i] / 2 + 1 / (2 * C_switch[i])
            print("满足约束条件的整体收益：")
            print(B)
            print("满足约束条件的整体收益个数：")
            print(len(B))
            if len(B) != 0:
                print("满足约束条件的整体收益的最大值max：")
                print(max(B.values()))

                # 存储系统效益
                wb = openpyxl.load_workbook('n5_Benefit.xlsx')
                ws = wb['Sheet']

                ws.cell(row=var + 1, column=1).value = var
                ws.cell(row=var + 1, column=2).value = max(B.values())

                wb.save('n5_Benefit.xlsx')

                print("最佳调度方案：")
                Plan_B.append(a.get((get_dict_key(B, max(B.values())))))
                print(Plan_B)
                _E = [data_new[Plan_B[0][0] - 1], data_new[Plan_B[0][1] - 1], data_new[Plan_B[0][2] - 1], data_new[Plan_B[0][3] - 1], data_new[Plan_B[0][4] - 1]]
                E = np.array(_E)
            else:
                print("无满足约束条件的调度方案，按原方案调度！")


# 根据字典的值value获得该值对应的key
def get_dict_key(dic, value):
    key = list(dic.keys())[list(dic.values()).index(value)]
    return key


if __name__ == "__main__":
    main()
