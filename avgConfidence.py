import sys

import numpy as np
import pandas as pd
import openpyxl
import random


def main():
    _data_r = pd.read_excel('a_r1.xlsx', usecols=['序号', 'd1', 'd2', 'd3', '当前状态', "选中次数", '置信度'])
    data_r = np.array(_data_r)
    _data_g = pd.read_excel('a_g1.xlsx', usecols=['序号', 'd1', 'd2', 'd3', '当前状态', "选中次数", '置信度'])
    data_g = np.array(_data_g)
    _data_b = pd.read_excel('a_b1.xlsx', usecols=['序号', 'd1', 'd2', 'd3', '当前状态', "选中次数", '置信度'])
    data_b = np.array(_data_b)
    _data_c = pd.read_excel('a_c1.xlsx', usecols=['序号', 'd1', 'd2', 'd3', '当前状态', "选中次数", '置信度'])
    data_c = np.array(_data_c)
    _data_y = pd.read_excel('a_y1.xlsx', usecols=['序号', 'd1', 'd2', 'd3', '当前状态', "选中次数", '置信度'])
    data_y = np.array(_data_y)

    # data_r[0][6] = 10
    # [h, l] = data_r.shape
    # print(h)
    # print(l)
    # print(np.array(data_r))
    # save(np.array(data_r), "a_r1.xlsx")
    # print(data_r[0][3])
    # print(data_g)
    # print(data_b)
    # print(data_c)
    # print(data_y)

    E1 = data_r[0]
    E2 = data_g[0]
    E3 = data_b[0]
    E4 = data_c[0]
    E5 = data_y[0]

    # print(E1)
    # print(E2)
    # print(E3)
    # print(E4)
    # print(E5)

    _E = []
    _E.append(E1)
    _E.append(E2)
    _E.append(E3)
    _E.append(E4)
    _E.append(E5)
    E = np.array(_E)
    print("选取五个执行体池中的第一个异构体")
    print("序号", "代码相异值", "传输相异值", "运行相异值", "当前状态", "选中次数", '置信度')
    print(E)

    # 按最后一列逆序排序
    # E_sort = E[np.lexsort(-E.T)]
    E_sort = np.array(sorted(E, key=lambda x: x[6], reverse=True))
    print("按置信度排序后:")
    print("序号", "代码相异值", "传输相异值", "运行相异值", "当前状态", "选中次数", '置信度')
    print(E_sort)

    print("从五个异构体中选取三个（置信度高优先）组成服务体S")
    print("服务体S：")
    print("序号", "代码相异值", "传输相异值", "运行相异值", "当前状态", "选中次数", '置信度')
    # 服务体S
    S = []
    # 选取的服务体的个数
    count = 0
    for i in range(1, 5):
        temp = E_sort[0][6]
        if E_sort[i][6] == temp:
            count += 1
    # count 0 1 2 3 4
    if count == 4:
        r1 = random.sample(range(0, 5), 3)
        S.append(E_sort[r1[0]])
        S[0][5] += 1
        S.append(E_sort[r1[1]])
        S[1][5] += 1
        S.append(E_sort[r1[2]])
        S[2][5] += 1
    elif count == 3:
        r1 = random.sample(range(0, 4), 3)
        S.append(E_sort[r1[0]])
        S[0][5] += 1
        S.append(E_sort[r1[1]])
        S[1][5] += 1
        S.append(E_sort[r1[2]])
        S[2][5] += 1
    elif count == 2:
        S.append(E_sort[0])
        S[0][5] += 1
        S.append(E_sort[1])
        S[1][5] += 1
        S.append(E_sort[2])
        S[2][5] += 1
    elif count == 1:
        S.append(E_sort[0])
        S[0][5] += 1
        S.append(E_sort[1])
        S[1][5] += 1
        count1 = 0
        for i in range(3, 5):
            temp = E_sort[2][6]
            if E_sort[i][6] == temp:
                count1 += 1
        if count1 == 0:
            S.append(E_sort[2])
            S[2][5] += 1
        elif count1 == 1:
            random1 = random.randint(2, 3)
            S.append(E_sort[random1])
            S[2][5] += 1
        else:
            random2 = random.randint(2, 4)
            S.append(E_sort[random2])
            S[2][5] += 1
    else:
        S.append(E_sort[0])
        S[0][5] += 1
        count2 = 0
        for i in range(2, 5):
            temp = E_sort[1][6]
            if E_sort[i][6] == temp:
                count2 += 1
        if count2 == 0:
            S.append(E_sort[1])
            S[1][5] += 1
            count3 = 0
            for i in range(3, 5):
                temp = E_sort[2][6]
                if E_sort[i][6] == temp:
                    count3 += 1
            if count3 == 0:
                S.append(E_sort[2])
                S[2][5] += 1
            elif count3 == 1:
                random3 = random.randint(2, 3)
                S.append(E_sort[random3])
                S[2][5] += 1
            else:
                random4 = random.randint(2, 4)
                S.append(E_sort[random4])
                S[2][5] += 1
        elif count2 == 1:
            S.append(E_sort[1])
            S[1][5] += 1
            S.append(E_sort[2])
            S[2][5] += 1
        elif count2 == 2:
            res1 = random.sample(range(1, 4), 2)
            S.append(E_sort[res1[0]])
            S[1][5] += 1
            S.append(E_sort[res1[1]])
            S[2][5] += 1
        else:
            res2 = random.sample(range(1, 5), 2)
            S.append(E_sort[res2[0]])
            S[1][5] += 1
            S.append(E_sort[res2[1]])
            S[2][5] += 1

    print(np.array(S))

    ids = [S[0][0], S[1][0], S[2][0]]

    for i in range(1, 5):
        var = E_sort[0][5]
        if var < E_sort[i][5]:
            var = E_sort[i][5]

    _var = int(var)
    print("第" + str(_var) + "次迭代！")
    input_req = input("请输入您需要的请求(操作1：1 操作2：2):")

    if int(input_req) == 1:
        print("进行操作1！")
    elif int(input_req) == 2:
        print("进行操作2！")
    else:
        print("输入错误，请重新输入!")
        sys.exit(0)

    for i in range(0, 3):
        if S[i][4] == 0:
            print("服务体S" + str(i) + "状态为" + str(S[i][4]) + ",输出正确！")
            # 更新置信度
            S[i][6] = ((S[i][5] * S[i][6]) + 1.0) / (S[i][5] + 1.0)
        elif S[i][4] == 3:
            print("服务体S" + str(i) + "状态为" + str(S[i][4]) + ",输出错误！")
            # S[i][6] = ((S[i][5] * S[i][6]) + 0.0) / (S[i][5] + 1.0)
            S[i][6] = 0
        elif S[i][4] == float(input_req):
            print("服务体S" + str(i) + "状态为" + str(S[i][4]) + ",输出错误！")
            # S[i][6] = ((S[i][5] * S[i][6]) + 0.0) / (S[i][5] + 1.0)
            S[i][6] = 0
        else:
            print("服务体S" + str(i) + "状态为" + str(S[i][4]) + ",输出正确！")
            S[i][6] = ((S[i][5] * S[i][6]) + 1.0) / (S[i][5] + 1.0)

    print("更新后的服务体S：")
    print("序号", "代码相异值", "传输相异值", "运行相异值", "当前状态", "选中次数", '置信度')
    print(np.array(S))

    H = 0
    for i in range(0, 3):
        H += S[i][6]
    print("初始平均置信度H(t):1.5")
    print("第" + str(_var) + "次迭代后，服务体的平均信誉度H(t)：" + str(H))

    # 触发清洗
    # for i in range(0, 5):
    #     if E[i][6] <= 0.125:
    #         print("触发清洗！")

    # 存储平均置信度
    wb = openpyxl.load_workbook('confidence.xlsx')
    ws = wb['Sheet']

    ws.cell(row=_var+1, column=1).value = var
    ws.cell(row=_var+1, column=2).value = H

    wb.save('confidence.xlsx')

    # 迭代更新xlsx
    # 获取工作簿，获取指定的sheet
    wb = openpyxl.load_workbook('a_r1.xlsx')
    ws = wb['Sheet']

    for rowNum in range(2, ws.max_row + 1):
        id = ws.cell(row=rowNum, column=1).value
        for i in range(0, len(ids)):
            if id == ids[i]:
                ws.cell(row=rowNum, column=10).value = S[i][5]
                ws.cell(row=rowNum, column=11).value = S[i][6]
            else:
                i += 1

    wb.save('a_r1.xlsx')

    wb = openpyxl.load_workbook('a_g1.xlsx')
    ws = wb['Sheet']

    for rowNum in range(2, ws.max_row + 1):
        id = ws.cell(row=rowNum, column=1).value
        for i in range(0, len(ids)):
            if id == ids[i]:
                ws.cell(row=rowNum, column=10).value = S[i][5]
                ws.cell(row=rowNum, column=11).value = S[i][6]
            else:
                i += 1

    wb.save('a_g1.xlsx')

    wb = openpyxl.load_workbook('a_b1.xlsx')
    ws = wb['Sheet']

    for rowNum in range(2, ws.max_row + 1):
        id = ws.cell(row=rowNum, column=1).value
        for i in range(0, len(ids)):
            if id == ids[i]:
                ws.cell(row=rowNum, column=10).value = S[i][5]
                ws.cell(row=rowNum, column=11).value = S[i][6]
            else:
                i += 1

    wb.save('a_b1.xlsx')

    wb = openpyxl.load_workbook('a_c1.xlsx')
    ws = wb['Sheet']

    for rowNum in range(2, ws.max_row + 1):
        id = ws.cell(row=rowNum, column=1).value
        for i in range(0, len(ids)):
            if id == ids[i]:
                ws.cell(row=rowNum, column=10).value = S[i][5]
                ws.cell(row=rowNum, column=11).value = S[i][6]
            else:
                i += 1

    wb.save('a_c1.xlsx')

    wb = openpyxl.load_workbook('a_y1.xlsx')
    ws = wb['Sheet']

    for rowNum in range(2, ws.max_row + 1):
        id = ws.cell(row=rowNum, column=1).value
        for i in range(0, len(ids)):
            if id == ids[i]:
                ws.cell(row=rowNum, column=10).value = S[i][5]
                ws.cell(row=rowNum, column=11).value = S[i][6]
            else:
                i += 1

    wb.save('a_y1.xlsx')


if __name__ == "__main__":
    main()
