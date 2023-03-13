import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
import warnings
warnings.filterwarnings('ignore')


# 正规化数据集 X
# def normalize(X, axis=-1, p=2):
#     lp_norm = np.atleast_1d(np.linalg.norm(X, p, axis))
#     lp_norm[lp_norm == 0] = 1
#     return X / np.expand_dims(lp_norm, axis)


# 计算一个样本与数据集中所有样本的欧氏距离的平方
def euclidean_distance(one_sample, X):
    one_sample = one_sample.reshape(1, -1)
    X = X.reshape(X.shape[0], -1)
    distances = np.power(np.tile(one_sample, (X.shape[0], 1)) - X, 2).sum(axis=1)
    return distances


# 保存数据到EXCEL
def save(data, path):
    wb = Workbook()
    ws = wb.active  # 激活 worksheet
    ws.append(['序号', '代码层面', 'd1', '传输层面', 'd2', '运行层面', 'd3', '异构度', '当前状态', "选中次数", '置信度'])

    [h, l] = data.shape  # h为行数，l为列数
    for i in range(h):
        row = []
        for j in range(l):
            row.append(data[i, j])
        ws.append(row)
    for i in range(h):
        ws.cell(i + 2, 10, 0)
        ws.cell(i + 2, 11, 0.5)
    wb.save(path)


class Kmeans():
    """Kmeans聚类算法.

    Parameters:
    -----------
    k: int
        聚类的数目.
    max_iterations: int
        最大迭代次数.
    varepsilon: float
        判断是否收敛, 如果上一次的所有k个聚类中心与本次的所有k个聚类中心的差都小于varepsilon,
        则说明算法已经收敛
    """

    def __init__(self, k=2, max_iterations=800, varepsilon=0.0001):
        self.k = k
        self.max_iterations = max_iterations
        self.varepsilon = varepsilon

    # 从所有样本中随机选取self.k样本作为初始的聚类中心
    def init_random_centroids(self, X):
        n_samples, n_features = np.shape(X)
        centroids = np.zeros((self.k, n_features))
        for i in range(self.k):
            centroid = X[np.random.choice(range(n_samples))]
            centroids[i] = centroid
        return centroids

    # 返回距离该样本最近的一个中心索引[0, self.k)
    def _closest_centroid(self, sample, centroids):
        distances = euclidean_distance(sample, centroids)
        closest_i = np.argmin(distances)
        return closest_i

    # 将所有样本进行归类，归类规则就是将该样本归类到与其最近的中心
    def create_clusters(self, centroids, X):
        n_samples = np.shape(X)[0]
        clusters = [[] for _ in range(self.k)]
        for sample_i, sample in enumerate(X):
            centroid_i = self._closest_centroid(sample, centroids)
            clusters[centroid_i].append(sample_i)
        return clusters

    # 对中心进行更新
    def update_centroids(self, clusters, X):
        n_features = np.shape(X)[1]
        centroids = np.zeros((self.k, n_features))
        for i, cluster in enumerate(clusters):
            centroid = np.mean(X[cluster], axis=0)
            centroids[i] = centroid
        return centroids

    # 将所有样本进行归类，其所在的类别的索引就是其类别标签
    def get_cluster_labels(self, clusters, X):
        y_pred = np.zeros(np.shape(X)[0])
        for cluster_i, cluster in enumerate(clusters):
            for sample_i in cluster:
                y_pred[sample_i] = cluster_i
        return y_pred

    # 对整个数据集X进行Kmeans聚类，返回其聚类的标签
    def predict(self, X):
        # 从所有样本中随机选取self.k样本作为初始的聚类中心
        centroids = self.init_random_centroids(X)

        # 迭代，直到算法收敛(上一次的聚类中心和这一次的聚类中心几乎重合)或者达到最大迭代次数
        for _ in range(self.max_iterations):
            # 将所有进行归类，归类规则就是将该样本归类到与其最近的中心
            clusters = self.create_clusters(centroids, X)
            former_centroids = centroids

            # 计算新的聚类中心
            centroids = self.update_centroids(clusters, X)

            # 如果聚类中心几乎没有变化，说明算法已经收敛，退出迭代
            diff = centroids - former_centroids
            if diff.any() < self.varepsilon:
                break

        return self.get_cluster_labels(clusters, X)


def main():
    # 1处理重复数据
    # 1.1读取excel中的数据；
    frame = pd.DataFrame(pd.read_excel('C:/Users/user/Desktop/test.xls', '异构度生成'))

    # 1.2去除重复行后的数据;
    no_re_row = frame.drop_duplicates()

    # 1.3保存去重后的数据到excel;
    no_re_row.to_excel('C:/Users/user/Desktop/test2.xlsx')

    data_no_re = pd.read_excel('C:/Users/user/Desktop/test2.xlsx')

    # Load the dataset
    _data = pd.read_excel('C:/Users/user/Desktop/test2.xlsx', usecols=['d1', 'd2', 'd3'])
    data = np.array(_data)
    num, dim = data.shape

    # 存放聚类完的五组数据
    a_r = []
    a_g = []
    a_b = []
    a_c = []
    a_y = []

    a_r1 = []
    a_g1 = []
    a_b1 = []
    a_c1 = []
    a_y1 = []

    # 用Kmeans算法进行聚类
    K = 5
    clf = Kmeans(K)
    y_pred = clf.predict(data)

    # 可视化聚类效果

    # 解决中文显示问题
    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['axes.unicode_minus'] = False

    color = ['r', 'g', 'b', 'c', 'y', 'm', 'k']
    ax = plt.subplot(111, projection='3d')
    ax.set_xlabel('代码层面')
    ax.set_ylabel('传输层面')
    ax.set_zlabel('运行层面')

    for p in range(0, num):
        y = y_pred[p]

        ax.scatter(data[p, 0], data[p, 1], data[p, 2], c=color[int(y)])
        item = color[int(y)]
        if item == 'r':
            a_r.append(p)
        elif item == 'g':
            a_g.append(p)
        elif item == 'b':
            a_b.append(p)
        elif item == 'c':
            a_c.append(p)
        elif item == 'y':
            a_y.append(p)
        else:
            print("出现漏网之鱼")

    # print(len(arr_r)+len(arr_g)+len(arr_b)+len(arr_c)+len(arr_y))
    # print(arr_r[0])
    # print(arr_r)
    # print(arr_g)
    # print(arr_b)
    # print(arr_c)
    # print(arr_y)

    print("聚类完成！")
    print("分成以下"+str(K)+"组：")
    print(a_r)
    print(a_g)
    print(a_b)
    print(a_c)
    print(a_y)

    # 根据p划分成的数组进行分组

    for i in range(0, len(a_r)):
        a_r1.append(data_no_re.iloc[a_r[i]])
    for i in range(0, len(a_g)):
        a_g1.append(data_no_re.iloc[a_g[i]])
    for i in range(0, len(a_b)):
        a_b1.append(data_no_re.iloc[a_b[i]])
    for i in range(0, len(a_c)):
        a_c1.append(data_no_re.iloc[a_c[i]])
    for i in range(0, len(a_y)):
        a_y1.append(data_no_re.iloc[a_y[i]])

    save(np.array(a_r1), "a_r1.xlsx")
    save(np.array(a_g1), "a_g1.xlsx")
    save(np.array(a_b1), "a_b1.xlsx")
    save(np.array(a_c1), "a_c1.xlsx")
    save(np.array(a_y1), "a_y1.xlsx")

    plt.show()

    wb = Workbook()
    ws = wb.active  # 激活 worksheet
    ws.cell(1, 1, 0)
    ws.cell(1, 2, 1.5)
    wb.save("confidence.xlsx")

if __name__ == "__main__":
    main()
